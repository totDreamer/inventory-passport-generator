from copy import copy

from openpyxl import load_workbook

from generators.inventory_fields import REPORT_FIELDS
from workflow_context import InventoryContext


# ==========================================================
# Константы
# ==========================================================

FIRST_DATA_ROW = 17
TEMPLATE_ROW = 17
OBJECT_NAME_CELL = "A8"
OBJECT_ADDRESS_CELL = "A10"
STAFF_COUNT_CELL = "A12"


# ==========================================================
# Главная функция
# ==========================================================

def generate_inventory_report(
    df,
    context: InventoryContext,
    output_path,
    start_row,
    end_row,
):
    """
    Генерирует инвентаризационный отчет.
    """

    try:

        workbook = open_template(context.workflow.inventory_template)
        sheet = workbook.active

        rows = select_rows(
            df,
            start_row,
            end_row,
        )

        if rows.empty:
            raise ValueError(
                "В выбранном диапазоне нет записей для создания отчета."
            )

        fill_inventory_context(sheet, context)

        fill_sheet(
            sheet,
            rows,
        )

        finalize_sheet(sheet)

        save_report(
            workbook,
            output_path,
        )

        return (
            True,
            f"Инвентаризационный отчет сохранен:\n{output_path}",
        )

    except Exception as e:

        return (
            False,
            f"Ошибка генерации отчета:\n{e}",
        )


# ==========================================================
# Работа с книгой
# ==========================================================

def open_template(template_path):
    """
    Открывает шаблон Excel.
    """

    return load_workbook(template_path)


def save_report(workbook, output_path):
    """
    Сохраняет отчет.
    """

    workbook.save(output_path)


# ==========================================================
# Работа с данными
# ==========================================================

def select_rows(
    df,
    start_row,
    end_row,
):
    """
    Возвращает выбранные пользователем строки.
    """

    return df.iloc[start_row - 1:end_row]


# ==========================================================
# Работа со строками шаблона
# ==========================================================

def insert_data_row(
    sheet,
    row_number,
    row_template,
):
    """
    Вставляет новую строку
    и полностью копирует оформление строки шаблона.
    """

    copy_row_style(
        sheet,
        row_number,
        row_template,
    )

    copy_row_dimensions(
        sheet,
        row_number,
        row_template,
    )


def copy_row_style(
    sheet,
    target_row,
    row_template,
):
    """
    Копирует стили строки.
    """

    for column, style in row_template["styles"].items():
        sheet.cell(target_row, column)._style = copy(style)


def copy_row_dimensions(
    sheet,
    target_row,
    row_template,
):
    """
    Копирует высоту строки.
    """

    template_dimension = row_template["dimension"]
    target_dimension = sheet.row_dimensions[target_row]

    for attribute in (
        "height",
        "hidden",
        "outlineLevel",
        "collapsed",
        "thickTop",
        "thickBot",
    ):
        setattr(
            target_dimension,
            attribute,
            copy(getattr(template_dimension, attribute)),
        )


def capture_row_template(sheet, row_number):
    """Capture cell styles and row dimensions before rows are inserted."""

    return {
        "styles": {
            cell.column: copy(cell._style)
            for cell in sheet[row_number]
        },
        "dimension": copy(sheet.row_dimensions[row_number]),
    }


def insert_data_rows(sheet, row_number, amount, row_template):
    """Insert styled rows and shift template metadata below the table."""

    merged_ranges = list(sheet.merged_cells.ranges)
    shifted_merges = []

    for merged in merged_ranges:
        if merged.max_row < row_number:
            continue

        sheet.unmerge_cells(str(merged))

        if merged.min_row >= row_number:
            min_row = merged.min_row + amount
            max_row = merged.max_row + amount
        else:
            min_row = merged.min_row
            max_row = merged.max_row + amount

        shifted_merges.append(
            (
                min_row,
                merged.min_col,
                max_row,
                merged.max_col,
            )
        )

    shifted_dimensions = {
        index: copy(dimension)
        for index, dimension in sheet.row_dimensions.items()
        if index >= row_number
    }

    for index in shifted_dimensions:
        del sheet.row_dimensions[index]

    sheet.insert_rows(row_number, amount=amount)

    for index, dimension in shifted_dimensions.items():
        target_index = index + amount
        dimension.index = target_index
        sheet.row_dimensions[target_index] = dimension

    for min_row, min_col, max_row, max_col in shifted_merges:
        sheet.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )

    for target_row in range(row_number, row_number + amount):
        insert_data_row(sheet, target_row, row_template)

# ==========================================================
# Заполнение листа
# ==========================================================

def fill_sheet(sheet, rows):
    """
    Заполняет отчет данными.
    """

    row_template = capture_row_template(sheet, TEMPLATE_ROW)
    insert_data_rows(
        sheet,
        FIRST_DATA_ROW,
        len(rows),
        row_template,
    )

    current_row = FIRST_DATA_ROW
    current_number = 1

    for data in rows.itertuples(index=False):

        fill_row(
            sheet=sheet,
            sheet_row=current_row,
            number=current_number,
            data=data,
        )

        current_row += 1
        current_number += 1


def fill_row(
    sheet,
    sheet_row,
    number,
    data,
):
    """
    Заполняет одну строку отчета.
    """

    write_cell(
        sheet,
        f"A{sheet_row}",
        number,
    )

    for column, builder in REPORT_FIELDS:

        value = builder(data)

        write_cell(
            sheet,
            f"{column}{sheet_row}",
            value,
        )


def write_cell(
    sheet,
    cell,
    value,
):
    """
    Записывает значение в ячейку.
    """

    if value is None:
        value = ""

    sheet[cell] = value


def fill_inventory_context(sheet, context):
    """Write report-level values into merged header anchors."""

    sheet[OBJECT_NAME_CELL] = (
        f"Наименование объекта: {context.object_name}"
    )
    sheet[OBJECT_ADDRESS_CELL] = f"Адрес: {context.object_address}"
    sheet[STAFF_COUNT_CELL] = (
        "Количество сотрудников (штат) на объекте: "
        f"{context.staff_count}"
    )


# ==========================================================
# Финальная обработка
# ==========================================================

def finalize_sheet(sheet):
    """
    Здесь можно будет разместить финальные действия
    перед сохранением книги.
    """

    pass
